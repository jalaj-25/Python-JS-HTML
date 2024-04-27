import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl

available_products = {
    1001: {"name": "Fruits", "price": 230, "category": "grocery", "quantity": 10, "date": "15/01/2024"},
    1002: {"name": "lotion", "price": 250, "category": "beauty & personal", "quantity": 100, "date": "15/01/2024"},
    1003: {"name": "combiflame", "price": 500, "category": "health", "quantity": 200, "date": "15/01/2024"},
    1004: {"name": "chilli sauce", "price": 20, "category": "grocery", "quantity": 50, "date": "15/01/2024"},
    1005: {"name": "toothbrush", "price": 700, "category": "beauty & personal", "quantity": 100, "date": "15/01/2024"},
    1006: {"name": "icecreams", "price": 33, "category": "grocery", "quantity": 56, "date": "15/01/2024"},
    1007: {"name": "nailpolish", "price": 765, "category": "beauty & personal", "quantity": 70, "date": "15/01/2024"},
    1008: {"name": "vegetables", "price": 764, "category": "grocery", "quantity": 90, "date": "15/01/2024"},
    1009: {"name": "facewash", "price": 87, "category": "beauty & personal", "quantity": 50, "date": "15/01/2024"},
    1010: {"name": "chocolate bars", "price": 30, "category": "grocery", "quantity": 60, "date": "15/01/2024"}
}

def display_data(data):
    for tree_entry in tree.get_children():
        tree.delete(tree_entry)
    
    for id, product in data.items():
        tree.insert("", tk.END, values=(id, product['name'], product['price'], product['category'], product['quantity'], product['date']))

def purchase_product():
    product_id = int(product_id_entry.get())
    if product_id in available_products:
        quantity = int(quantity_entry.get())
        if available_products[product_id]['quantity'] >= quantity:
            available_products[product_id]['quantity'] -= quantity
            display_data(available_products)
            save_purchase_to_excel(product_id, quantity, name_entry.get(), contact_entry.get())
            messagebox.showinfo("Success", "Purchase successful!")
        else:
            messagebox.showerror("Error", "Insufficient quantity!")
    else:
        messagebox.showerror("Error", "Invalid Product ID!")

def save_purchase_to_excel(product_id, quantity, customer_name, contact_number):
    try:
        wb = openpyxl.load_workbook("purchase_data.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Product ID", "Name", "Price", "Category", "Quantity", "Date", "Customer Name", "Contact Number"])
    else:
        ws = wb.active

    product = available_products[product_id]
    ws.append([product_id, product['name'], product['price'], product['category'], quantity, product['date'], customer_name, contact_number])
    wb.save("purchase_data.xlsx")
    messagebox.showinfo("Success", "Purchase data saved to Excel file.")

root = tk.Tk()
root.title("Inventory Management System")

frame = ttk.Frame(root)
frame.grid(row=0, column=0, padx=10, pady=10)

tree = ttk.Treeview(frame, columns=("ID", "Name", "Price", "Category", "Quantity", "Date"), show="headings")
tree.heading("ID", text="ID")
tree.heading("Name", text="Name")
tree.heading("Price", text="Price")
tree.heading("Category", text="Category")
tree.heading("Quantity", text="Quantity")
tree.heading("Date", text="Date")
tree.grid(row=0, column=0, columnspan=3, pady=10)

display_data(available_products)

product_id_label = ttk.Label(frame, text="Product ID:")
product_id_label.grid(row=1, column=0, padx=5)
product_id_entry = ttk.Entry(frame)
product_id_entry.grid(row=1, column=1, padx=5)

name_label = ttk.Label(frame, text="Name:")
name_label.grid(row=2, column=0, padx=5)
name_entry = ttk.Entry(frame)
name_entry.grid(row=2, column=1, padx=5)

contact_label = ttk.Label(frame, text="Contact Number:")
contact_label.grid(row=3, column=0, padx=5)
contact_entry = ttk.Entry(frame)
contact_entry.grid(row=3, column=1, padx=5)

quantity_label = ttk.Label(frame, text="Quantity:")
quantity_label.grid(row=4, column=0, padx=5)
quantity_entry = ttk.Entry(frame)
quantity_entry.grid(row=4, column=1, padx=5)

purchase_button = ttk.Button(frame, text="Purchase", command=purchase_product)
purchase_button.grid(row=5, column=1, pady=10)

root.mainloop()

































# import openpyxl
# from tkinter import messagebox  # Import messagebox directly from tkinter for error handling
# import tkinter as tk  # Import tk for use with Treeview and other widgets
# from tkinter import ttk  # Import ttk for themed widgets

# available_products = {
#     1001: {"name": "Fruits", "price": 230, "category": "grocery", "quantity": 10, "date": "15/01/2024"},
#     1002: {"name": "lotion", "price": 250, "category": "beauty & personal", "quantity": 100, "date": "15/01/2024"},
#     1003: {"name": "combiflame", "price": 500, "category": "health", "quantity": 200, "date": "15/01/2024"},
#     1004: {"name": "chilli sauce", "price": 20, "category": "grocery", "quantity": 50, "date": "15/01/2024"},
#     1005: {"name": "toothbrush", "price": 700, "category": "beauty & personal", "quantity": 100, "date": "15/01/2024"},
#     1006: {"name": "icecreams", "price": 33, "category": "grocery", "quantity": 56, "date": "15/01/2024"},
#     1007: {"name": "nailpolish", "price": 765, "category": "beauty & personal", "quantity": 70, "date": "15/01/2024"},
#     1008: {"name": "vegetables", "price": 764, "category": "grocery", "quantity": 90, "date": "15/01/2024"},
#     1009: {"name": "facewash", "price": 87, "category": "beauty & personal", "quantity": 50, "date": "15/01/2024"},
#     1010: {"name": "chocolate bars", "price": 30, "category": "grocery", "quantity": 60, "date": "15/01/2024"}
# } 

# def display_data(tree):
#     for item in tree.get_children():
#         tree.delete(item)

#     for id, product in available_products.items():
#         tree.insert("", tk.END, values=(id, product['name'], product['price'], product['category'], product['quantity'], product['date']))

# def purchase_product(product_id, quantity, name, contact, tree):
#     global available_products  # Use the global available_products dictionary

#     if product_id in available_products:
#         if available_products[product_id]['quantity'] >= quantity:
#             available_products[product_id]['quantity'] -= quantity
            
#             new_quantity = available_products[product_id]['quantity']
#             if new_quantity <= 10:
#                 messagebox.showinfo("Low Stock", f"Product ID {product_id} is now low in stock! Quantity: {new_quantity}")

#             update_tree_item(tree, product_id, new_quantity)

#             save_purchase_to_excel(product_id, quantity, name, contact)
#             return True
#         else:
#             messagebox.showerror("Error", "Insufficient quantity!")
#             return False
#     else:
#         messagebox.showerror("Error", "Invalid Product ID!")
#         return False

# def update_tree_item(tree, product_id, new_quantity):
#     for item in tree.get_children():
#         values = tree.item(item, 'values')
#         if values[0] == product_id:
#             tree.item(item, values=(values[0], values[1], values[2], values[3], new_quantity, values[5]))
#             break

# def save_purchase_to_excel(product_id, quantity, customer_name, contact_number):
#     try:
#         wb = openpyxl.load_workbook("purchase_data.xlsx")
#     except FileNotFoundError:
#         wb = openpyxl.Workbook()
#         ws = wb.active 
#         ws.append(["Product ID", "Name", "Price", "Category", "Quantity", "Date", "Customer Name", "Contact Number"])
#     else:
#         ws = wb.active

#     try:
#         product = available_products[product_id]
#         ws.append([product_id, product['name'], product['price'], product['category'], quantity, product['date'], customer_name, contact_number])
#         wb.save("purchase_data.xlsx")
#         print(f"Purchase saved successfully to purchase_data.xlsx")
#     except Exception as e:
#         print(f"Failed to save purchase to purchase_data.xlsx: {e}")
#         messagebox.showerror("Error", f"Failed to save purchase to Excel file: {e}")